from openpyxl import load_workbook

class ExcelReader:
    
    @staticmethod
    def getTestData(path, sheetName):
        worksheet = load_workbook(path)
        sheet = worksheet[sheetName]
        
        data = []
        
        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            data.append(row_data) 
                
        return data